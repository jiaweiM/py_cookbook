from openpyxl import Workbook
from openpyxl import load_workbook
import pytest


def test_create():
    wb = Workbook()
    ws = wb.active
    ws1 = wb.create_sheet("Mysheet") # insert at the end (default)
    ws2 = wb.create_sheet("Mysheet", 0)
    ws.title = 'New Title'


def test_load():
    wb2 = load_workbook('test.xlsx')
